"""
Excel report builder for structural risk (riesgo estructural).
Produces a workbook with four sheets showing step-by-step scoring.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styling helpers ──────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TITLE_FONT = Font(bold=True, name="Arial", size=13, color="2F5496")
SECTION_FONT = Font(bold=True, name="Arial", size=10, color="2F5496")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(bold=True, name="Arial", size=10)
SCORE_FONT = Font(bold=True, name="Arial", size=14, color="2F5496")
SCORE_LABEL_FONT = Font(bold=True, name="Arial", size=12)
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
WARNING_FONT = Font(bold=True, name="Arial", size=10, color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _cell(ws, r, c, val=None, font=NORMAL_FONT, fmt=None, fill=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = font
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell


def _header_row(ws, r, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 50)


# ── Sheet builders ───────────────────────────────────────────────

def build_concentracion_sheet(wb, data: dict):
    ws = wb.create_sheet("Concentración")
    conc = data["concentracion"]
    r = 1

    ws.cell(row=r, column=1, value="RIESGO DE CONCENTRACIÓN").font = TITLE_FONT
    r += 2

    # 1. Weights per subyacente
    ws.cell(row=r, column=1, value="1. Pesos por subyacente (wi)").font = SECTION_FONT
    r += 1
    _header_row(ws, r, ["Subyacente", "Monto (USD)", "Peso (wi)", "wi²"])
    r += 1
    for item in conc["items"]:
        _cell(ws, r, 1, item["nombre"])
        _cell(ws, r, 2, item["monto"], fmt="#,##0.00")
        _cell(ws, r, 3, item["weight"], fmt="0.0000%")
        _cell(ws, r, 4, item["weight_2"], fmt="0.000000")
        r += 1
    _cell(ws, r, 1, "TOTAL", font=BOLD_FONT)
    _cell(ws, r, 2, sum(i["monto"] for i in conc["items"]), font=BOLD_FONT, fmt="#,##0.00")
    _cell(ws, r, 3, sum(i["weight"] for i in conc["items"]), font=BOLD_FONT, fmt="0.0000%")
    _cell(ws, r, 4, conc["hhi"], font=BOLD_FONT, fmt="0.000000")
    r += 2

    # 2. HHI result
    ws.cell(row=r, column=1, value="2. Índice HHI").font = SECTION_FONT
    r += 1
    _cell(ws, r, 1, "HHI = Σ wi²")
    _cell(ws, r, 2, conc["hhi"], fmt="0.000000")
    r += 1
    _cell(ws, r, 1, "Inversiones equivalentes (1/HHI)")
    _cell(ws, r, 2, conc["inversiones_totales"], fmt="0.0000")
    r += 2

    # 3. Max weight
    ws.cell(row=r, column=1, value="3. Peso de la mayor posición").font = SECTION_FONT
    r += 1
    _cell(ws, r, 1, "Mayor posición")
    _cell(ws, r, 2, conc["max_weight_nombre"])
    r += 1
    _cell(ws, r, 1, "Peso")
    _cell(ws, r, 2, conc["max_weight"], fmt="0.00%")
    r += 2
    _header_row(ws, r, ["Mayor posición", "Score", "Interpretación"])
    r += 1
    for rng, s, interp in [
        ("≤ 15%", 10, "Concentración muy sana"),
        ("15% – 25%", 8, "Aceptable"),
        ("25% – 35%", 6, "Concentración visible"),
        ("35% – 45%", 4, "Posición dominante relevante"),
        ("45% – 60%", 2, "Dependencia excesiva"),
        ("> 60%", 1, "Riesgo crítico"),
    ]:
        _cell(ws, r, 1, rng)
        _cell(ws, r, 2, s)
        _cell(ws, r, 3, interp)
        r += 1
    _cell(ws, r, 1, "Score peso máximo", font=BOLD_FONT)
    _cell(ws, r, 2, conc["max_weight_score"], font=BOLD_FONT, fill=YELLOW_FILL)
    _cell(ws, r, 3, conc.get("max_weight_interpretacion", ""), font=BOLD_FONT)
    r += 2

    # 4. Blended indicator → HHI table lookup
    ws.cell(row=r, column=1, value="4. Score final de concentración").font = SECTION_FONT
    r += 1
    _cell(ws, r, 1, "Fórmula")
    _cell(ws, r, 2, "total = 0.70 × HHI + 0.30 × peso máximo")
    r += 1
    hhi_val = conc["hhi"]
    mw_val = conc["max_weight"]
    _cell(ws, r, 1, f"0.70 × HHI (0.70 × {hhi_val:.6f})")
    _cell(ws, r, 2, round(0.7 * hhi_val, 6), fmt="0.000000")
    r += 1
    _cell(ws, r, 1, f"0.30 × peso máx (0.30 × {mw_val:.6f})")
    _cell(ws, r, 2, round(0.3 * mw_val, 6), fmt="0.000000")
    r += 1
    total = conc.get("total", round(0.7 * hhi_val + 0.3 * mw_val, 6))
    _cell(ws, r, 1, "Total", font=BOLD_FONT)
    _cell(ws, r, 2, total, font=BOLD_FONT, fmt="0.000000")
    r += 2

    # HHI table applied to total
    _cell(ws, r, 1, "Búsqueda en tabla HHI", font=BOLD_FONT)
    r += 1
    _header_row(ws, r, ["HHI", "Score", "Interpretación"])
    r += 1
    for rng, s, interp in [
        ("≤ 0.13", "9–10", "Excelente diversificación"),
        ("0.13 – 0.17", "8", "Muy bien diversificado"),
        ("0.17 – 0.22", "7", "Bien diversificado"),
        ("0.22 – 0.28", "6", "Diversificación razonable"),
        ("0.28 – 0.34", "5", "Concentración moderada"),
        ("0.34 – 0.41", "4", "Concentrado"),
        ("0.41 – 0.51", "3", "Muy concentrado"),
        ("0.51 – 0.66", "2", "Riesgo alto"),
        ("> 0.66", "1", "Riesgo crítico"),
    ]:
        _cell(ws, r, 1, rng)
        _cell(ws, r, 2, s)
        _cell(ws, r, 3, interp)
        r += 1
    r += 1

    _cell(ws, r, 1, "SCORE CONCENTRACIÓN", font=SCORE_LABEL_FONT)
    _cell(ws, r, 2, conc["score"], font=SCORE_FONT, fill=YELLOW_FILL, fmt="0.00")
    _cell(ws, r, 3, conc.get("interpretacion", ""), font=BOLD_FONT)

    _auto_width(ws)


def build_correlacion_sheet(wb, data: dict):
    ws = wb.create_sheet("Correlación")
    corr = data["correlacion"]
    r = 1

    ws.cell(row=r, column=1, value="RIESGO DE CORRELACIÓN").font = TITLE_FONT
    r += 2

    # 1. Weights
    ws.cell(row=r, column=1, value="1. Pesos por clase de activo").font = SECTION_FONT
    r += 1
    _header_row(ws, r, ["Clase de activo", "Peso (%)"])
    r += 1
    for name, w in corr["subyacentes_weights"].items():
        _cell(ws, r, 1, name)
        _cell(ws, r, 2, w / 100, fmt="0.00%")
        r += 1
    r += 1

    # 2. Matrices
    assets = list(corr["subyacentes_weights"].keys())
    matrix_vals = {
        (c["asset_i"], c["asset_j"]): c for c in corr["correlation_matrix"]
    }

    ws.cell(row=r, column=1, value="2. Matriz de correlación ponderada (wi × wj × ρij)").font = SECTION_FONT
    r += 1

    # ρij
    ws.cell(row=r, column=1, value="Coeficientes de correlación (ρij)").font = BOLD_FONT
    r += 1
    short = [a[:20] for a in assets]
    _header_row(ws, r, [""] + short)
    r += 1
    for ai in assets:
        _cell(ws, r, 1, ai[:20], font=BOLD_FONT)
        for j, aj in enumerate(assets):
            val = matrix_vals.get((ai, aj), {})
            _cell(ws, r, j + 2, val.get("corr", 0), fmt="0.00")
        r += 1
    r += 1

    # wi × wj × ρij
    ws.cell(row=r, column=1, value="Valores ponderados (wi × wj × ρij)").font = BOLD_FONT
    r += 1
    _header_row(ws, r, [""] + short)
    r += 1
    for ai in assets:
        _cell(ws, r, 1, ai[:20], font=BOLD_FONT)
        for j, aj in enumerate(assets):
            val = matrix_vals.get((ai, aj), {})
            _cell(ws, r, j + 2, val.get("value", 0), fmt="0.000000")
        r += 1
    r += 1

    # 3. Total
    ws.cell(row=r, column=1, value="3. Correlación ponderada total").font = SECTION_FONT
    r += 1
    _cell(ws, r, 1, "Σ (wi × wj × ρij)")
    _cell(ws, r, 2, corr["total_correlation"], fmt="0.000000")
    r += 2

    # 4. Score table
    ws.cell(row=r, column=1, value="4. Escala de Score").font = SECTION_FONT
    r += 1
    _header_row(ws, r, ["Correlación promedio", "Score", "Interpretación"])
    r += 1
    for rng, s, interp in [
        ("≤ 0.25", 10, "Diversificación excepcional"),
        ("0.26 – 0.35", "8–9", "Diversificación real (multi-factor)"),
        ("0.36 – 0.45", 7, "Portafolio eficiente"),
        ("0.46 – 0.52", 6, "Correlación moderada"),
        ("0.53 – 0.57", 5, "Riesgo sistémico creciente"),
        ("0.58 – 0.62", 4, "Alta sincronización"),
        ("0.63 – 0.68", 3, "Riesgo estructural alto"),
        ("> 0.68", "1–2", "Efecto dominó"),
    ]:
        _cell(ws, r, 1, rng)
        _cell(ws, r, 2, str(s))
        _cell(ws, r, 3, interp)
        r += 1
    r += 1

    _cell(ws, r, 1, "SCORE CORRELACIÓN", font=SCORE_LABEL_FONT)
    _cell(ws, r, 2, corr["score"], font=SCORE_FONT, fill=YELLOW_FILL)
    _cell(ws, r, 3, corr.get("interpretacion", ""), font=BOLD_FONT)

    _auto_width(ws)


def build_entity_sheet(wb, data: dict):
    ws = wb.create_sheet("Gestor y Administrador")
    r = 1

    ws.cell(row=r, column=1, value="CALIDAD DE GESTOR Y ADMINISTRADOR").font = TITLE_FONT
    r += 2

    for section_name, key, label in [
        ("1. Score por Gestor", "gestor", "Gestor"),
        ("2. Score por Administrador", "administrador", "Administrador"),
    ]:
        entity = data[key]
        ws.cell(row=r, column=1, value=section_name).font = SECTION_FONT
        r += 1
        ws.cell(row=r, column=1, value="Fórmula: Score = Σ (puntos_i × peso_i)").font = NORMAL_FONT
        r += 1
        _header_row(ws, r, [label, "Peso (wi)", "Puntos", "Aporte (puntos × wi)"])
        r += 1
        for g in entity["groups"]:
            _cell(ws, r, 1, g["name"])
            _cell(ws, r, 2, g["weight"], fmt="0.00%")
            _cell(ws, r, 3, g["points"], fmt="0.0")
            _cell(ws, r, 4, g["weighted_points"], fmt="0.000000")
            r += 1
        _cell(ws, r, 1, "TOTAL", font=BOLD_FONT)
        _cell(ws, r, 2, sum(g["weight"] for g in entity["groups"]), font=BOLD_FONT, fmt="0.00%")
        _cell(ws, r, 3, "", font=BOLD_FONT)
        _cell(ws, r, 4, entity["score"], font=BOLD_FONT, fmt="0.000000")
        r += 1

        _cell(ws, r, 1, f"SCORE {label.upper()}", font=SCORE_LABEL_FONT)
        _cell(ws, r, 2, round(entity["score"], 2), font=SCORE_FONT, fill=YELLOW_FILL, fmt="0.00")
        r += 2

        if entity.get("unmatched"):
            ws.cell(row=r, column=1, value="⚠ Entidades no encontradas en tabla de referencia").font = WARNING_FONT
            r += 1
            for u in entity["unmatched"]:
                _cell(ws, r, 1, u["value"])
                _cell(ws, r, 2, u["weight"], fmt="0.00%")
                r += 1
        r += 1

    # Contribution to global
    ws.cell(row=r, column=1, value="3. Contribución al score global").font = SECTION_FONT
    r += 1
    _header_row(ws, r, ["Dimensión", "Score", "Peso global", "Aporte"])
    r += 1
    for lbl, key, peso in [("Gestor", "gestor", 0.15), ("Administrador", "administrador", 0.15)]:
        s = data[key]["score"]
        _cell(ws, r, 1, lbl)
        _cell(ws, r, 2, round(s, 2), fmt="0.00")
        _cell(ws, r, 3, peso, fmt="0%")
        _cell(ws, r, 4, round(s * peso, 4), fmt="0.0000")
        r += 1

    _auto_width(ws)


def build_moneda_sheet(wb, data: dict):
    ws = wb.create_sheet("Moneda")
    mon = data["moneda"]
    r = 1

    ws.cell(row=r, column=1, value="EXPOSICIÓN CAMBIARIA").font = TITLE_FONT
    r += 2

    # 1. PEN exposure
    ws.cell(row=r, column=1, value="1. Exposición a PEN").font = SECTION_FONT
    r += 1
    _cell(ws, r, 1, "Total en PEN (USD equiv.)")
    _cell(ws, r, 2, mon["pen_total"], fmt="#,##0.00")
    r += 1
    _cell(ws, r, 1, "% del portafolio en PEN")
    _cell(ws, r, 2, mon["pen_pct"], fmt="0.00%")
    r += 2

    # 2. Score table
    ws.cell(row=r, column=1, value="2. Escala de Score").font = SECTION_FONT
    r += 1
    _header_row(ws, r, ["Exposición a PEN", "Score"])
    r += 1
    for rng, s in [
        ("≤ 10%", 10), ("10 – 15%", 9), ("15 – 20%", 8),
        ("20 – 25%", 7), ("25 – 30%", 6), ("30 – 35%", 5),
        ("35 – 40%", 4), ("40 – 45%", 3), ("45 – 55%", 2), ("> 55%", 1),
    ]:
        _cell(ws, r, 1, rng)
        _cell(ws, r, 2, s)
        r += 1
    r += 1

    _cell(ws, r, 1, "SCORE MONEDA", font=SCORE_LABEL_FONT)
    _cell(ws, r, 2, mon["score"], font=SCORE_FONT, fill=YELLOW_FILL)
    r += 2

    # 3. Global score summary
    ws.cell(row=r, column=1, value="3. Score Global de Riesgo Estructural").font = SECTION_FONT
    r += 1
    _header_row(ws, r, ["Dimensión", "Score", "Peso", "Aporte"])
    r += 1
    dims = [
        ("Concentración", data["concentracion"]["score"], 0.25),
        ("Correlación", data["correlacion"]["score"], 0.25),
        ("Moneda", data["moneda"]["score"], 0.20),
        ("Gestor", data["gestor"]["score"], 0.15),
        ("Administrador", data["administrador"]["score"], 0.15),
    ]
    for lbl, s, w in dims:
        _cell(ws, r, 1, lbl)
        _cell(ws, r, 2, round(s, 2), fmt="0.00")
        _cell(ws, r, 3, w, fmt="0%")
        _cell(ws, r, 4, round(s * w, 4), fmt="0.0000")
        r += 1
    _cell(ws, r, 1, "SCORE GLOBAL", font=SCORE_LABEL_FONT)
    _cell(ws, r, 2, "", font=BOLD_FONT)
    _cell(ws, r, 3, "", font=BOLD_FONT)
    _cell(ws, r, 4, data["global_score"], font=SCORE_FONT, fill=YELLOW_FILL, fmt="0.00")

    _auto_width(ws)


# ── Public API ───────────────────────────────────────────────────

def build_structural_risk_excel(data: dict) -> Workbook:
    """
    Build a complete structural risk Excel workbook from scoring data.

    Args:
        data: Dict with keys: concentracion, correlacion, gestor,
              administrador, moneda, global_score.

    Returns:
        openpyxl.Workbook ready to save or stream.
    """
    wb = Workbook()
    wb.remove(wb.active)

    build_concentracion_sheet(wb, data)
    build_correlacion_sheet(wb, data)
    build_entity_sheet(wb, data)
    build_moneda_sheet(wb, data)

    return wb
