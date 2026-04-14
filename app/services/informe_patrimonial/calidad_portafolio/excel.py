"""
FastAPI service that uses Anthropic SDK to generate alignment analysis
and produces an Excel file with three sheets.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import ReportRequest


# ── Styling helpers ──────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TITLE_FONT = Font(bold=True, name="Arial", size=12, color="2F5496")
SECTION_FONT = Font(bold=True, name="Arial", size=10, color="2F5496")
NORMAL_FONT = Font(name="Arial", size=10)
SCORE_FONT = Font(bold=True, name="Arial", size=14, color="2F5496")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
LIGHT_BLUE_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


# ── Sheet builders ───────────────────────────────────────────────

def build_asset_sheet(wb, req: ReportRequest):
    ws = wb.create_sheet("Alineacion Activo")
    d = req.asset_alignment
    r = 1

    # Title
    ws.cell(row=r, column=1, value="ALINEACIÓN POR TIPO DE ACTIVO").font = TITLE_FONT
    r += 2

    # Section 1: Universe
    ws.cell(row=r, column=1, value="1. Universo del análisis").font = SECTION_FONT
    r += 1
    for label, val in [
        ("Inversionista", req.inversionista.title()),
        ("Patrimonio invertible", f"USD {req.total_patrimonio:,.2f}"),
        ("Perfil de riesgo", req.perfil_riesgo),
        ("Benchmark aplicable",
         f"{'≥500K' if req.total_patrimonio >= 500000 else '<500K'} – Perfil {req.perfil_riesgo}"),
    ]:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=val).font = NORMAL_FONT
        r += 1
    r += 1

    # Section 2: Benchmark
    ws.cell(row=r, column=1, value="2. Benchmark utilizado").font = SECTION_FONT
    r += 1
    headers = ["Clase de activo", "Benchmark %"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    for ad in d.asset_details:
        style_data_cell(ws, r, 1).value = ad.asset_name
        style_data_cell(ws, r, 2, "0.00%").value = ad.benchmark_percentage / 100
        r += 1
    ws.cell(row=r, column=1, value="Total").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=r, column=2, value=1).number_format = "0.00%"
    ws.cell(row=r, column=2).font = Font(bold=True, name="Arial", size=10)
    r += 2

    # Section 3: Ranges
    ws.cell(row=r, column=1, value="3. Rangos permitidos por clase de activo").font = SECTION_FONT
    r += 1
    headers = ["Clase de activo", "Límite inferior", "Límite superior"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    for ad in d.asset_details:
        style_data_cell(ws, r, 1).value = ad.asset_name
        style_data_cell(ws, r, 2, "0.00%").value = ad.min_limit / 100
        style_data_cell(ws, r, 3, "0.00%").value = ad.max_limit / 100
        r += 1
    r += 1

    # Section 4: Comparison & penalties
    ws.cell(row=r, column=1, value="4. Comparación vs benchmark y penalizaciones").font = SECTION_FONT
    r += 1
    headers = ["Clase de activo", "Benchmark", "Rango Min", "Rango Max", req.inversionista.title() or "Inversionista", "Desviación",
               "Penalización", "¿Dentro?"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    for ad in d.asset_details:
        style_data_cell(ws, r, 1).value = ad.asset_name
        style_data_cell(ws, r, 2, "0.00%").value = ad.benchmark_percentage / 100
        style_data_cell(ws, r, 3, "0.00%").value = ad.min_limit / 100
        style_data_cell(ws, r, 4, "0.00%").value = ad.max_limit / 100
        style_data_cell(ws, r, 5, "0.00%").value = ad.investor_percentage / 100
        style_data_cell(ws, r, 6, "0.00%").value = ad.deviation / 100
        style_data_cell(ws, r, 7, "0.00%").value = ad.penalty / 100
        cell = style_data_cell(ws, r, 8)
        cell.value = "Sí" if ad.is_within_limits else "No"
        cell.fill = GREEN_FILL if ad.is_within_limits else RED_FILL
        r += 1
    # Totals row
    ws.cell(row=r, column=1, value="TOTAL PENALIZACIÓN").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=r, column=7, value=d.penalizacion_total / 100).number_format = "0.00%"
    ws.cell(row=r, column=7).font = Font(bold=True, name="Arial", size=10)
    r += 2

    # Section 5: Structural distance & score
    ws.cell(row=r, column=1, value="5. Distancia estructural y Score").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Distancia estructural").font = NORMAL_FONT
    ws.cell(row=r, column=2, value=d.distancia_estructural / 100).number_format = "0.00%"
    r += 1
    ws.cell(row=r, column=1, value="Score de Alineación por Tipo de Activo").font = Font(bold=True, name="Arial",
                                                                                         size=11)
    cell = ws.cell(row=r, column=2, value=f"{d.score} / 10")
    cell.font = SCORE_FONT
    cell.fill = YELLOW_FILL

    auto_width(ws)


def build_risk_sheet(wb, req: ReportRequest):
    ws = wb.create_sheet("Alineacion Riesgo")
    d = req.risk_alignment
    r = 1

    ws.cell(row=r, column=1, value="ALINEACIÓN POR RIESGO").font = TITLE_FONT
    r += 2

    # Investment contributions table
    ws.cell(row=r, column=1, value="1. Score de performance ponderado por inversión").font = SECTION_FONT
    r += 1
    headers = ["Producto", "Tipo de activo", "Monto (USD)", "Peso (%)", "Risk Score", "Aporte"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    for ic in d.investment_contributions:
        style_data_cell(ws, r, 1).value = ic.producto_nombre
        style_data_cell(ws, r, 2).value = ", ".join(ic.tipo_activo)
        style_data_cell(ws, r, 3, "#,##0.00").value = ic.monto
        style_data_cell(ws, r, 4, "0.00%").value = ic.peso_inversion / 100
        style_data_cell(ws, r, 5, "0.00").value = ic.risk_score
        style_data_cell(ws, r, 6, "0.0000").value = ic.aporte
        r += 1
    # Total
    ws.cell(row=r, column=1, value="TOTAL PONDERADO").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=r, column=6, value=d.score_total_weighted).font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=r, column=6).number_format = "0.0000"
    r += 2

    # Profile range
    ws.cell(row=r, column=1, value="2. Rango objetivo según perfil").font = SECTION_FONT
    r += 1
    profiles = [
        ("Conservador", "7 – 8", "6.5 – 9.0"),
        ("Conservador & Moderado", "6 – 7", "5.5 – 8.0"),
        ("Moderado", "5 – 6", "4.5 – 6.5"),
        ("Moderado & Agresivo", "4 – 5", "3.5 – 5.5"),
        ("Agresivo", "3 – 4", "2.0 – 4.5"),
    ]
    headers = ["Perfil", "Rango objetivo", "Rango tolerado"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    for prof, rng, tol in profiles:
        c = style_data_cell(ws, r, 1)
        c.value = prof
        if prof == d.perfil_riesgo:
            c.fill = YELLOW_FILL
            style_data_cell(ws, r, 2).fill = YELLOW_FILL
            style_data_cell(ws, r, 3).fill = YELLOW_FILL
        style_data_cell(ws, r, 2).value = rng
        style_data_cell(ws, r, 3).value = tol
        r += 1
    r += 1

    # D calculation explanation
    ws.cell(row=r, column=1, value="3. Cálculo de distancia al rango (d)").font = SECTION_FONT
    r += 1
    dc = d.d_calculation
    for label, val in [
        ("Score total ponderado", f"{dc.score_total:.4f}"),
        ("Rango perfil", f"[{dc.perfil_min}, {dc.perfil_max}]"),
        ("Zona central", f"[{dc.first_quarter}, {dc.third_quarter}]"),
        ("Zona borde inferior", f"[{dc.perfil_min}, {dc.first_quarter})"),
        ("Zona borde superior", f"({dc.third_quarter}, {dc.perfil_max}]"),
        ("Zona del score", dc.zone.replace("_", " ").title()),
        ("Punto de referencia", f"{dc.reference_point}"),
        ("d value", f"{dc.d_value}"),
    ]:
        ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=val).font = NORMAL_FONT
        r += 1
    r += 1

    # Score translation table
    ws.cell(row=r, column=1, value="4. Traducción a Score (1-10)").font = SECTION_FONT
    r += 1
    headers = ["Condición", "Score"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    score_table = [
        ("Zona central [L+0.25, U-0.25]", 10),
        ("Zona borde [L, L+0.25) o (U-0.25, U]", 9),
        ("d ∈ (0.00, 0.25]", 8),
        ("d ∈ (0.25, 0.50]", 7),
        ("d ∈ (0.50, 0.75]", 6),
        ("d ∈ (0.75, 1.00]", 5),
        ("d ∈ (1.00, 1.50]", 4),
        ("d ∈ (1.50, 2.00]", 2),
        ("d > 2.00", 1),
    ]
    for cond, sc in score_table:
        style_data_cell(ws, r, 1).value = cond
        style_data_cell(ws, r, 2).value = sc
        r += 1
    r += 1

    # Final result
    ws.cell(row=r, column=1, value="Score de Alineación por Riesgo").font = Font(bold=True, name="Arial", size=11)
    cell = ws.cell(row=r, column=2, value=f"{d.score} / 10")
    cell.font = SCORE_FONT
    cell.fill = YELLOW_FILL
    r += 1
    ws.cell(row=r, column=1, value="Interpretación").font = NORMAL_FONT
    ws.cell(row=r, column=2, value=getattr(d, 'interpretation', '')).font = NORMAL_FONT
    r += 1
    ws.cell(row=r, column=1, value="Zona de interpretación").font = NORMAL_FONT
    zone_label = getattr(d, 'interpretation_zone', '').replace("_", " ").title()
    ws.cell(row=r, column=2, value=zone_label).font = NORMAL_FONT

    auto_width(ws)


def build_geo_sheet(wb, req: ReportRequest):
    ws = wb.create_sheet("Alineacion Geografica")
    d = req.geo_alignment
    r = 1

    ws.cell(row=r, column=1, value="ALINEACIÓN GEOGRÁFICA").font = TITLE_FONT
    r += 2

    # Benchmark
    ws.cell(row=r, column=1, value="1. Benchmark utilizado (Sabbi Cracks)").font = SECTION_FONT
    r += 1
    headers = ["Región", "Benchmark"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    for rd in d.region_details:
        style_data_cell(ws, r, 1).value = rd.region.title()
        style_data_cell(ws, r, 2, "0.00%").value = rd.benchmark_percentage / 100
        r += 1
    r += 1

    # Ranges
    ws.cell(row=r, column=1, value="2. Rangos permitidos por región").font = SECTION_FONT
    r += 1
    headers = ["Región", "Tolerancia", "Límite inferior", "Límite superior"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    for rd in d.region_details:
        style_data_cell(ws, r, 1).value = rd.region.title()
        style_data_cell(ws, r, 2).value = rd.tolerance
        style_data_cell(ws, r, 3, "0.00%").value = rd.min_limit / 100
        style_data_cell(ws, r, 4, "0.00%").value = rd.max_limit / 100
        r += 1
    r += 1

    # Comparison
    ws.cell(row=r, column=1, value="3. Distribución geográfica del portafolio").font = SECTION_FONT
    r += 1
    headers = ["Región", "Benchmark", "Lím. Inf.", "Lím. Sup.", req.inversionista.title(), "Desviación", "Penalización",
               "¿Dentro?"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    for rd in d.region_details:
        style_data_cell(ws, r, 1).value = rd.region.title()
        style_data_cell(ws, r, 2, "0.00%").value = rd.benchmark_percentage / 100
        style_data_cell(ws, r, 3, "0.00%").value = rd.min_limit / 100
        style_data_cell(ws, r, 4, "0.00%").value = rd.max_limit / 100
        style_data_cell(ws, r, 5, "0.00%").value = rd.portfolio_percentage / 100
        style_data_cell(ws, r, 6, "0.00%").value = rd.deviation / 100
        style_data_cell(ws, r, 7, "0.00%").value = rd.penalty / 100
        cell = style_data_cell(ws, r, 8)
        cell.value = "Sí" if rd.is_within_limits else "No"
        cell.fill = GREEN_FILL if rd.is_within_limits else RED_FILL
        r += 1
    r += 1

    # Score table
    ws.cell(row=r, column=1, value="4. Traducción a Score de Concentración Geográfica").font = SECTION_FONT
    r += 1
    headers = ["Desviación total", "Score", "Interpretación"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    score_table = [
        ("0 – 5%", 10, "Muy bien diversificado"),
        ("5 – 10%", 9, "Diversificación sólida"),
        ("10 – 15%", 8, "Desvío menor"),
        ("15 – 20%", 7, "Desvío moderado"),
        ("20 – 25%", 6, "Riesgo relevante"),
        ("25 – 30%", 5, "Concentración importante"),
        ("30 – 35%", 4, "Alta concentración"),
        ("35 – 45%", 3, "Riesgo elevado"),
        ("45 – 60%", 2, "Riesgo crítico"),
        ("> 60%", 1, "Riesgo extremo"),
    ]
    for dev, sc, interp in score_table:
        style_data_cell(ws, r, 1).value = dev
        style_data_cell(ws, r, 2).value = sc
        style_data_cell(ws, r, 3).value = interp
        r += 1
    r += 1

    # Final result
    ws.cell(row=r, column=1, value="Desviación geográfica total (bruta)").font = NORMAL_FONT
    ws.cell(row=r, column=2, value=d.total_deviation / 100).number_format = "0.00%"
    r += 1
    structural_dev = getattr(d, 'structural_deviation', d.total_deviation / 2)
    ws.cell(row=r, column=1, value="Desviación estructural (ajustada ÷2)").font = NORMAL_FONT
    ws.cell(row=r, column=2, value=structural_dev / 100).number_format = "0.00%"
    r += 1
    ws.cell(row=r, column=1, value=f"Interpretación").font = NORMAL_FONT
    ws.cell(row=r, column=2, value=d.interpretation).font = NORMAL_FONT
    r += 1
    ws.cell(row=r, column=1, value="Score de Concentración Geográfica").font = Font(bold=True, name="Arial", size=11)
    cell = ws.cell(row=r, column=2, value=f"{d.score} / 10")
    cell.font = SCORE_FONT
    cell.fill = YELLOW_FILL

    auto_width(ws)
