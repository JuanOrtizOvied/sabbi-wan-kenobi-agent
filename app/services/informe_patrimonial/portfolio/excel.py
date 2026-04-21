import io
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.api.informe_patrimonial.portfolio.schemas import (
    PortafolioItem,
)

# ──────────────────────────────────────────────
# Styles
# ──────────────────────────────────────────────

_FONT = Font(name="Arial", size=10)
_BOLD_FONT = Font(name="Arial", size=10, bold=True)

_HEADER_FILL = PatternFill("solid", fgColor="F0F0F0")
_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

_ALIGN = Alignment(horizontal="left", vertical="center")
_ALIGN_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

_CURRENCY_FMT = "#,##0.00"
_PCT_FMT = "0.00%"

# ──────────────────────────────────────────────
# Column layout
# ──────────────────────────────────────────────
# Flat fields: columns 1–13
_FLAT_HEADERS = [
    "cuenta_bancaria_inversion",  # A
    "tipo_activo",                # B
    "pertenencia",                # C
    "moneda_invertida",           # D
    "valor_estimado_usd",         # E
    "rendimiento_anual_porcentaje",  # F
    "name",                       # G
    "slugs",                      # H
    "comision_sin_igv",           # I
    "moneda",                     # J
    "administrador",              # K
    "gestor",                     # L
    "liquidez",                   # M
]
_NUM_FLAT = len(_FLAT_HEADERS)  # 13

# Nested groups: 3 columns each starting after flat fields
# clase_activo  → cols 14-16 (N-P)
# foco_geografico → cols 17-19 (Q-S)
# subyacente    → cols 20-22 (T-V)
_NESTED_GROUPS = [
    {"label": "clase_activo", "attr": "clase_activo", "start_col": _NUM_FLAT + 1},
    {"label": "foco_geografico", "attr": "foco_geografico", "start_col": _NUM_FLAT + 4},
    {"label": "subyacente", "attr": "subyacente", "start_col": _NUM_FLAT + 7},
]
_SUB_HEADERS = ["name", "percentage", "slugs"]

_TOTAL_COLS = _NUM_FLAT + 9  # 22

_COL_WIDTHS = {
    "A": 48, "B": 20, "C": 12, "D": 14, "E": 18,
    "F": 16, "G": 42, "H": 48, "I": 14, "J": 12,
    "K": 20, "L": 20, "M": 16,
    "N": 30, "O": 12, "P": 38,
    "Q": 22, "R": 12, "S": 22,
    "T": 36, "U": 12, "V": 36,
}


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def _set_cell(ws, row, col, value, bold=False, fill=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _BOLD_FONT if bold else _FONT
    cell.alignment = _ALIGN
    cell.border = _BORDER
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    return cell


def _write_main_header(ws):
    """Row 1: flat headers + merged group labels for nested arrays."""
    # Flat headers
    for col_idx, label in enumerate(_FLAT_HEADERS, 1):
        _set_cell(ws, 1, col_idx, label, bold=True, fill=_HEADER_FILL)

    # Nested group headers (merged across 3 sub-columns each)
    for group in _NESTED_GROUPS:
        sc = group["start_col"]
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc + 2)
        _set_cell(ws, 1, sc, group["label"], bold=True, fill=_HEADER_FILL)
        for c in range(sc + 1, sc + 3):
            ws.cell(row=1, column=c).border = _BORDER
            ws.cell(row=1, column=c).fill = _HEADER_FILL


def _write_sub_header_row(ws, row):
    """Write the repeating sub-header row for nested groups (grey fill)."""
    for group in _NESTED_GROUPS:
        sc = group["start_col"]
        for i, label in enumerate(_SUB_HEADERS):
            _set_cell(ws, row, sc + i, label, bold=True, fill=_HEADER_FILL)


def _get_flat_values(item) -> list:
    """Extract the 13 flat-field values from a portfolio item."""
    slugs_str = ", ".join(item.slugs) if item.slugs else ""
    return [
        item.cuenta_bancaria_inversion,
        item.tipo_activo,
        item.pertenencia,
        item.moneda_invertida,
        item.valor_estimado_usd,
        item.rendimiento_anual_porcentaje,
        item.name,
        slugs_str,
        item.comision_sin_igv,
        item.moneda,
        item.administrador,
        item.gestor,
        item.liquidez,
    ]


def _write_item_block(ws, start_row, item) -> int:
    """
    Write one portfolio item block. Returns the number of rows used.

    Layout:
      Row 0 (sub-header): flat fields (merged) | name | percentage | slugs × 3
      Row 1..N (data):    (merged)              | data values...
    """
    n_clase = len(item.clase_activo) if item.clase_activo else 0
    n_foco = len(item.foco_geografico) if item.foco_geografico else 0
    n_sub = len(item.subyacente) if item.subyacente else 0
    data_rows = max(n_clase, n_foco, n_sub, 1)
    total_rows = 1 + data_rows  # sub-header + data

    # ── Sub-header row ──
    _write_sub_header_row(ws, start_row)

    # ── Flat fields on the sub-header row, merged down ──
    flat_values = _get_flat_values(item)
    for col_idx, value in enumerate(flat_values, 1):
        fmt = None
        if col_idx == 5:   # valor_estimado_usd
            fmt = _CURRENCY_FMT
        elif col_idx == 6:  # rendimiento_anual_porcentaje
            fmt = _PCT_FMT
        _set_cell(ws, start_row, col_idx, value, number_format=fmt)

    if total_rows > 1:
        for col_idx in range(1, _NUM_FLAT + 1):
            ws.merge_cells(
                start_row=start_row, start_column=col_idx,
                end_row=start_row + total_rows - 1, end_column=col_idx,
            )
    # Set wrap on long text fields
    for col_idx in (1, 7, 8):
        ws.cell(row=start_row, column=col_idx).alignment = _ALIGN_WRAP

    # ── Data rows for nested arrays ──
    for i in range(data_rows):
        r = start_row + 1 + i

        # clase_activo
        if item.clase_activo and i < len(item.clase_activo):
            ca = item.clase_activo[i]
            sc = _NESTED_GROUPS[0]["start_col"]
            _set_cell(ws, r, sc, ca.name)
            _set_cell(ws, r, sc + 1, ca.percentage)
            _set_cell(ws, r, sc + 2, ", ".join(ca.slugs))
        else:
            sc = _NESTED_GROUPS[0]["start_col"]
            for c in range(3):
                _set_cell(ws, r, sc + c, None)

        # foco_geografico
        if item.foco_geografico and i < len(item.foco_geografico):
            fg = item.foco_geografico[i]
            sc = _NESTED_GROUPS[1]["start_col"]
            _set_cell(ws, r, sc, fg.name)
            _set_cell(ws, r, sc + 1, fg.percentage)
            _set_cell(ws, r, sc + 2, ", ".join(fg.slugs))
        else:
            sc = _NESTED_GROUPS[1]["start_col"]
            for c in range(3):
                _set_cell(ws, r, sc + c, None)

        # subyacente
        if item.subyacente and i < len(item.subyacente):
            sb = item.subyacente[i]
            sc = _NESTED_GROUPS[2]["start_col"]
            _set_cell(ws, r, sc, sb.name)
            _set_cell(ws, r, sc + 1, sb.percentage)
            _set_cell(ws, r, sc + 2, ", ".join(sb.slugs))
        else:
            sc = _NESTED_GROUPS[2]["start_col"]
            for c in range(3):
                _set_cell(ws, r, sc + c, None)

    # Ensure borders on all cells in sub-header row
    for c in range(1, _TOTAL_COLS + 1):
        ws.cell(row=start_row, column=c).border = _BORDER

    return total_rows


# ──────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────

def generate_portfolio_excel(portafolio: List[PortafolioItem]) -> bytes:
    """
    Build a .xlsx workbook that replicates the nested-subtable layout:

      Row 1: main headers (flat fields + merged group labels)
      Per item:
        - sub-header row (name|percentage|slugs × 3) with grey fill
        - 1–N data rows
        - flat fields merged vertically across the block
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    _write_main_header(ws)

    current_row = 2
    for item in portafolio:
        rows_used = _write_item_block(ws, current_row, item)
        current_row += rows_used

    for letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
